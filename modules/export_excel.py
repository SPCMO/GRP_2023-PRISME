# -*- coding: utf-8 -*-
"""Export Excel des résultats de campagne — généré à la demande depuis results_store
(pas pendant le run), en plus du dashboard interactif (décision validée avec
l'utilisateur). 7 onglets, demandés explicitement pour retrouver hors de l'outil tout
ce qui est visible dans le Dashboard : Paramétrage (contexte de la campagne), Vue
synthèse, Détail par crue, Sensibilité au seuil, Vue 3D, Variation selon le nb de crues
— soit les 5 sous-onglets du Dashboard, un par un, plus le contexte de campagne — et
Analyse crues affl. (contribution des stations affluentes, indépendant de tout calage).

Les graphiques matplotlib sont RE-rendus ici (backend Agg, non interactif) plutôt que
réutilisés depuis ui/tab_dashboard.py ou ui/tab_analyse_affluents.py : ce module
(modules/) ne doit pas dépendre de ui/ (sens de dépendance imposé par l'architecture du
projet — voir le plan). Quelques petites constantes (couleurs, conversion
horizon->minutes) et 2 fonctions de calcul pur (barycentre de la pluie, percentiles du
temps de réponse) sont donc dupliquées volontairement plutôt que partagées entre les
deux couches — même doctrine déjà appliquée aux autres onglets de ce module.
"""

import os
import re
from datetime import timedelta
from io import BytesIO

import numpy as np
from matplotlib import dates as mdates
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401 — nécessaire pour projection="3d"
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from modules import affluents, results_store
from modules.criteres_perf import CriteresPerfError, parse_criteres_perf, parse_evenement_serie
from modules.grp_paths import construire_grp_paths
from modules.score import (
    calculer_scores, config_ponderation_par_defaut, explication_score, filtrer_par_crues,
    resoudre_ponderation,
)

_COULEUR_OBS = "#1B4F72"
_PALETTE_COURBES = (
    "#CC5500", "#1D6A39", "#7B241C", "#7D3C98", "#117864", "#B7950B",
    "#2874A6", "#A93226", "#5D6D7E", "#943126",
)
_MARQUEURS_METHODE = {"T": "o", "R": "h"}

# Mêmes libellés/couleurs que ui/tab_config.py::LIBELLES_SEUILS_Q — dupliqué pour la
# même raison de sens de dépendance (modules/ ne dépend pas de ui/).
LIBELLES_SEUILS_Q = (
    ("zt_jaune", "ZT Jaune"), ("jaune", "Jaune"),
    ("zt_orange", "ZT Orange"), ("orange", "Orange"),
    ("zt_rouge", "ZT Rouge"), ("rouge", "Rouge"),
)
# Version couleur (3-uplet), nécessaire pour "Analyse crues affl." — les autres feuilles
# n'ont besoin que du libellé (2-uplet ci-dessus), d'où les 2 constantes séparées.
_LIBELLES_SEUILS_Q_COULEUR = (
    ("zt_jaune", "ZT Jaune", "#9A7D0A"), ("jaune", "Jaune", "#9A7D0A"),
    ("zt_orange", "ZT Orange", "#784212"), ("orange", "Orange", "#784212"),
    ("zt_rouge", "ZT Rouge", "#641E16"), ("rouge", "Rouge", "#641E16"),
)
_COULEUR_LOCAL = "#BDC3C7"  # gris, écoulements locaux non expliqués par un affluent suivi

ENTETES_DETAIL = ("Crue", "Date/heure crue", "Horizon", "Seuil C1", "Méthode", "Statut",
                   "dQP (%)", "dTP (pdt)", "VE (%)", "KGE", "Suspect", "Note")
ENTETES_SYNTHESE = ("Horizon", "Seuil C1", "Méthode", "Score composite", "Nb crues",
                     "Moyenne |dQP| (%)", "Moyenne |dTP|", "Moyenne |VE| (%)", "Moyenne (1-KGE)")
ENTETES_SENSIBILITE = ("Horizon", "Seuil C1", "Méthode", "Score composite")
ENTETES_VUE3D = ("Horizon", "Seuil C1", "Méthode", "Score composite",
                  "Écart au meilleur (Δ)", "Moyenne |dQP| (%)", "Moyenne |dTP|",
                  "Moyenne |VE| (%)", "Moyenne (1-KGE)")
ENTETES_CRUES = ("Crue", "Date/heure de début", "Qmax observé (m³/s)",
                  "Cumul de pluie de l'épisode (mm)")
ENTETES_VARIATION = ("N crues (les plus fortes, Qmax décroissant)", "Combinaison gagnante",
                      "Score normalisé (à ce N)", "KGE moyen (brut)", "Moyenne |dQP| (brut, %)",
                      "Moyenne |dTP| (brut, pdt)")
ENTETES_AFFLUENTS_CONFIG = ("Nom", "Code station", "Surface BV (km²)",
                             "Propagation P10", "Propagation P50", "Propagation P90",
                             "Fichier de débits")
ENTETES_BILAN_AFFLUENTS = ("Station", "Surface BV (km²)", "Volume transité (hm³)",
                            "% du volume exutoire", "Q à Qmax exutoire (m³/s)",
                            "% du Qmax exutoire")

_MOTIF_HORIZON = re.compile(r"(\d{2})J(\d{2})H(\d{2})M")


def _horizon_en_minutes(horizon):
    """Même logique que ui.tab_dashboard._horizon_en_minutes — voir ce module pour le
    détail (conversion 'ddJhhHmmM' -> minutes, pour trier/positionner numériquement)."""
    m = _MOTIF_HORIZON.match(horizon or "")
    if not m:
        return 0
    j, h, mn = (int(g) for g in m.groups())
    return j * 1440 + h * 60 + mn


def _intervalle_minutes(serie):
    """Durée réelle (minutes) entre 2 points consécutifs d'une série EVxxxx.DAT — mesurée
    sur les horodatages plutôt que supposée depuis un code pas de temps. Ne sert qu'à
    dimensionner la LARGEUR des barres de l'hyétogramme (voir _figure_detail_crue) :
    Pobs est déjà en mm par pas de temps (l'en-tête "Pobs(mm/h)" du fichier est
    trompeur — confirmé par comparaison avec les cumuls réellement observés pour un
    événement majeur réel), donc la hauteur des barres n'a plus besoin de cette durée,
    voir _cumul_pluie_mm."""
    if len(serie) < 2:
        return None
    delta = (serie[1][0] - serie[0][0]).total_seconds() / 60
    return delta if delta > 0 else None


def _cumul_pluie_mm(serie):
    """Pobs est déjà en mm par pas de temps : somme directe, sans conversion (l'en-tête
    "Pobs(mm/h)" du fichier est trompeur — une interprétation en intensité mm/h avait
    d'abord été tentée mais donnait des cumuls trop faibles pour de vrais événements
    majeurs, confirmé par l'utilisateur sur la crue historique de l'Aude d'octobre 2018,
    Qmax=1648 m³/s : 33mm calculés contre 150-300mm réellement observés)."""
    if not serie:
        return None
    return sum(p[1] for p in serie)


def _fig_to_image(fig):
    """Rend une Figure matplotlib (backend Agg, non interactif, dpi déjà fixé à la
    création de chaque Figure) en openpyxl Image, sans fichier temporaire sur disque."""
    buf = BytesIO()
    FigureCanvasAgg(fig).print_png(buf)
    buf.seek(0)
    return XLImage(buf)


def _ajuster_largeurs(ws, largeurs):
    for i, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largeur


def _entete(ws, entetes, largeurs=None):
    ws.append(entetes)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    _ajuster_largeurs(ws, largeurs or [16] * len(entetes))


# ══════════════════════════════════════════════════════════════════════════════════
# Construction des infos par crue (communes aux onglets Paramétrage et Détail par crue)
# ══════════════════════════════════════════════════════════════════════════════════

def _construire_infos_crues(paths, app, dates_iso):
    """Pour chaque date de crue (ISO) déjà en base : cherche l'événement correspondant
    dans CRITERES_PERF.DAT (tous les pas de temps configurés, le premier qui matche
    l'emporte — en pratique un seul pas de temps est utilisé) pour son n° d'événement et
    son Qmax, puis lit sa série EVxxxx.DAT pour le cumul de pluie (mm, somme directe
    des valeurs de la série — voir _cumul_pluie_mm) et pour le tracé (onglet Détail par
    crue).
    Best-effort : une crue absente de CRITERES_PERF.DAT ou dont la série est illisible
    reste incluse, juste avec les champs correspondants à None — jamais une erreur
    bloquante pour l'export entier."""
    infos = {iso: {"num_evt": None, "code_pdt": None, "date_deb": None, "qmax_obs": None,
                    "cumul_pluie_mm": None, "serie": None}
              for iso in dates_iso}
    if paths is None:
        return infos

    pdt_list = app.config_data.get("parametrage", {}).get("pas_de_temps", [])
    restants = set(dates_iso)
    for pdt in pdt_list:
        if not restants:
            break
        code_pdt = pdt["code"]
        try:
            evenements = parse_criteres_perf(paths.criteres_perf_dat(code_pdt))
        except (FileNotFoundError, CriteresPerfError):
            continue
        for evt in evenements:
            iso = evt.date_deb.isoformat()
            if iso not in restants:
                continue
            entree = infos[iso]
            entree["num_evt"], entree["code_pdt"] = evt.num_evt, code_pdt
            entree["date_deb"], entree["qmax_obs"] = evt.date_deb, evt.qmax
            chemin_serie = os.path.join(paths.evenements_dir(code_pdt),
                                         f"{paths.code_site}-EV{evt.num_evt:04d}.DAT")
            try:
                serie = parse_evenement_serie(chemin_serie)
            except (FileNotFoundError, CriteresPerfError):
                serie = None
            entree["serie"] = serie
            entree["cumul_pluie_mm"] = _cumul_pluie_mm(serie) if serie else None
            restants.discard(iso)
    return infos


def _libelle_crue(iso, info):
    d = info.get("date_deb")
    if d is None:
        from datetime import datetime as _dt
        d = _dt.fromisoformat(iso)
    prefixe = f"#{info['num_evt']}" if info.get("num_evt") is not None else "?"
    return f"{prefixe} - {d:%d/%m/%Y %H:%M}"


# ══════════════════════════════════════════════════════════════════════════════════
# Onglet 1 — Paramétrage
# ══════════════════════════════════════════════════════════════════════════════════

def _feuille_parametrage(ws, app, couverture, infos_crues, poids, asymetrie_dtp, libelle_profil,
                           crues_incluses):
    station = app.config_data.get("station", {})
    seuils_q = app.config_data.get("seuils_q", {})
    pdt_list = app.config_data.get("parametrage", {}).get("pas_de_temps", [])

    def _titre(texte):
        ws.append((texte,))
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        ws.append(())

    _titre("Station")
    for libelle, cle in (("Code station", "code_station"), ("Code site (GRP)", "code_site"),
                          ("Nom de la station", "nom_station"), ("Code BNBV", "code_bnbv")):
        ws.append((libelle, station.get(cle) or "—"))
    ws.append(())

    _titre("Seuils de vigilance en débit (PHyC)")
    ws.append(("Niveau", "Seuil (m³/s)"))
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for cle, libelle in LIBELLES_SEUILS_Q:
        val = seuils_q.get(cle)
        ws.append((libelle, val if val is not None else "—"))
    ws.append(())

    _titre("Pas de temps de calage configuré(s)")
    ws.append(("Libellé", "Code GRP"))
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for pdt in pdt_list:
        ws.append((pdt.get("libelle"), pdt.get("code")))
    ws.append(())

    _titre("Horizons testés (nb de combinaisons réussies / tentées)")
    ws.append(("Horizon", "Réussies", "Tentées"))
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for h in sorted(couverture["horizons"], key=_horizon_en_minutes):
        c = couverture["horizons"][h]
        ws.append((h, c["complets"], c["tentes"]))
    ws.append(())

    _titre("Seuils de calage testés (nb de combinaisons réussies / tentées)")
    ws.append(("Seuil C1", "Réussies", "Tentées"))
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for s in sorted(couverture["seuils"]):
        c = couverture["seuils"][s]
        ws.append((s, c["complets"], c["tentes"]))
    ws.append(())

    _titre("Méthode(s) de correction de sortie testée(s) (nb de combinaisons réussies / tentées)")
    ws.append(("Méthode", "Réussies", "Tentées"))
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for m in sorted(couverture["methodes"]):
        c = couverture["methodes"][m]
        ws.append((("Tangara" if m == "T" else "RNA" if m == "R" else m), c["complets"], c["tentes"]))
    ws.append(())

    _titre("Pondération du score composite actuellement active : " + libelle_profil)
    ws.append(("Indicateur", "Poids"))
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for cle, libelle in (("dqp", "|dQP|"), ("dtp", "|dTP|"), ("ve", "|VE|"), ("kge", "(1−KGE)")):
        ws.append((libelle, poids.get(cle)))
    ws.append(("Asymétrie dTP — facteur retard (dTP > 0)", asymetrie_dtp.get("retard")))
    ws.append(("Asymétrie dTP — facteur avance (dTP < 0)", asymetrie_dtp.get("avance")))
    if crues_incluses:
        ws.append((f"Crues incluses dans le score : {len(crues_incluses)} sur "
                    f"{len(infos_crues)} disponibles (sélection restreinte, voir Dashboard "
                    "> \"Crues dans le score\")",))
    else:
        ws.append(("Crues incluses dans le score : toutes",))
    ws.append(())
    ws.append(("Explication détaillée du calcul :",))
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True)
    for ligne_txt in explication_score(poids, asymetrie_dtp).split("\n"):
        ws.append((ligne_txt,))
    ws.append(())

    _titre(f"Crues testées ({len(infos_crues)}) — voir l'onglet \"Détail par crue\" pour "
           "les indicateurs simulés par combinaison")
    ws.append(ENTETES_CRUES)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for iso, info in sorted(infos_crues.items(),
                             key=lambda kv: (kv[1]["num_evt"] is None, kv[1]["num_evt"] or 0)):
        ws.append((
            _libelle_crue(iso, info),
            info["date_deb"].strftime("%d/%m/%Y %H:%M") if info["date_deb"] else iso,
            round(info["qmax_obs"], 2) if info["qmax_obs"] is not None else "—",
            round(info["cumul_pluie_mm"], 1) if info["cumul_pluie_mm"] is not None else "—",
        ))

    _ajuster_largeurs(ws, [46, 20, 20, 20])


# ══════════════════════════════════════════════════════════════════════════════════
# Onglet 2 — Vue synthèse
# ══════════════════════════════════════════════════════════════════════════════════

def _feuille_vue_synthese(ws, lignes_ok, scores_valides, meilleur):
    # -- Classement complet (même contenu que le tableau "Classement" du Dashboard) --
    _entete(ws, ENTETES_SYNTHESE, [16, 12, 12, 18, 10, 18, 18, 18, 18])
    for s in scores_valides:
        ws.append((
            s.horizon, s.seuil_c1, s.methode, round(s.score, 4), s.nb_crues,
            round(s.moyennes_erreur.get("dqp"), 2) if s.moyennes_erreur.get("dqp") is not None else None,
            round(s.moyennes_erreur.get("dtp"), 2) if s.moyennes_erreur.get("dtp") is not None else None,
            round(s.moyennes_erreur.get("ve"), 2) if s.moyennes_erreur.get("ve") is not None else None,
            round(s.moyennes_erreur.get("kge"), 4) if s.moyennes_erreur.get("kge") is not None else None,
        ))

    ligne_libre = ws.max_row + 2
    horizons = sorted({s.horizon for s in scores_valides}, key=_horizon_en_minutes)
    seuils = sorted({s.seuil_c1 for s in scores_valides})

    # -- Heatmap horizon x seuil (score moyen, toutes méthodes confondues) — reproduite
    # en tableau avec dégradé de couleur Excel (ColorScaleRule) à la place du imshow
    # matplotlib de l'app : reste triable/filtrable, contrairement à une image. --
    if horizons and seuils:
        r0 = ligne_libre
        ws.cell(row=r0, column=1, value="Score composite moyen (0 = meilleur) — Seuil \\ Horizon").font = \
            Font(bold=True)
        for j, h in enumerate(horizons):
            ws.cell(row=r0 + 1, column=2 + j, value=h).font = Font(bold=True)
        grille = {}
        for i, s_val in enumerate(seuils):
            ws.cell(row=r0 + 2 + i, column=1, value=s_val).font = Font(bold=True)
        for s in scores_valides:
            i, j = seuils.index(s.seuil_c1), horizons.index(s.horizon)
            grille[(i, j)] = s.score if (i, j) not in grille else (grille[(i, j)] + s.score) / 2
        for (i, j), val in grille.items():
            ws.cell(row=r0 + 2 + i, column=2 + j, value=round(val, 4))
        plage = (f"{get_column_letter(2)}{r0 + 2}:"
                 f"{get_column_letter(1 + len(horizons))}{r0 + 1 + len(seuils)}")
        ws.conditional_formatting.add(plage, ColorScaleRule(
            start_type="min", start_color="1D6A39", mid_type="percentile", mid_value=50,
            mid_color="F9E79F", end_type="max", end_color="943126"))
        ligne_libre = r0 + 2 + len(seuils) + 2

    # -- Dispersion |dQP| par horizon : tableau min/Q1/médiane/Q3/max (équivalent texte
    # de la boîte à moustaches de l'app, qui n'a pas d'équivalent natif Excel). --
    if horizons:
        r0 = ligne_libre
        ws.cell(row=r0, column=1, value="Dispersion |dQP| (%) par horizon").font = Font(bold=True)
        entetes_disp = ("Horizon", "Min", "1er quartile", "Médiane", "3e quartile", "Max", "Nb valeurs")
        for j, e in enumerate(entetes_disp):
            ws.cell(row=r0 + 1, column=1 + j, value=e).font = Font(bold=True)
        valeurs_par_horizon = {}
        for h in horizons:
            valeurs_par_horizon[h] = [abs(l["dqp"]) for l in lignes_ok
                                        if l["horizon"] == h and l["dqp"] is not None]
        for i, h in enumerate(horizons):
            v = valeurs_par_horizon[h]
            if v:
                ligne = (h, round(min(v), 2), round(float(np.percentile(v, 25)), 2),
                          round(float(np.median(v)), 2), round(float(np.percentile(v, 75)), 2),
                          round(max(v), 2), len(v))
            else:
                ligne = (h, "—", "—", "—", "—", "—", 0)
            for j, val in enumerate(ligne):
                ws.cell(row=r0 + 2 + i, column=1 + j, value=val)
        ligne_libre = r0 + 2 + len(horizons) + 2

    # -- Image : reproduction du graphique heatmap + dispersion de l'app --
    if horizons and seuils:
        fig = _figure_vue_synthese(horizons, seuils, scores_valides, lignes_ok, meilleur)
        img = _fig_to_image(fig)
        ws.add_image(img, f"A{ligne_libre}")


def _figure_vue_synthese(horizons, seuils, scores_valides, lignes_ok, meilleur):
    fig = Figure(figsize=(11, 4.3), dpi=115)
    ax_heat = fig.add_subplot(1, 2, 1)
    ax_disp = fig.add_subplot(1, 2, 2)
    fig.subplots_adjust(wspace=0.4, bottom=0.2)

    grille = np.full((len(seuils), len(horizons)), np.nan)
    for s in scores_valides:
        i, j = seuils.index(s.seuil_c1), horizons.index(s.horizon)
        grille[i, j] = s.score if np.isnan(grille[i, j]) else (grille[i, j] + s.score) / 2
    im = ax_heat.imshow(grille, aspect="auto", cmap="RdYlGn_r", vmin=0, vmax=1)
    ax_heat.set_xticks(range(len(horizons)))
    ax_heat.set_xticklabels(horizons, rotation=45, ha="right", fontsize=7)
    ax_heat.set_yticks(range(len(seuils)))
    ax_heat.set_yticklabels([f"{v:.2f}" for v in seuils], fontsize=7)
    ax_heat.set_title("Score composite (0=meilleur)", fontsize=9)
    fig.colorbar(im, ax=ax_heat, fraction=0.046, pad=0.04)
    if meilleur is not None and meilleur.horizon in horizons and meilleur.seuil_c1 in seuils:
        from matplotlib.patches import Rectangle
        j, i = horizons.index(meilleur.horizon), seuils.index(meilleur.seuil_c1)
        ax_heat.add_patch(Rectangle((j - 0.5, i - 0.5), 1, 1, fill=False,
                                     edgecolor="#FFD700", linewidth=2.5, zorder=5))

    positions = [_horizon_en_minutes(h) for h in horizons]
    positions_triees = sorted(positions)
    ecarts = [b - a for a, b in zip(positions_triees, positions_triees[1:]) if b > a]
    largeur_boite = max(min(ecarts) * 0.4, 1) if ecarts else 1
    valeurs_par_horizon = []
    for h in horizons:
        valeurs = [abs(l["dqp"]) for l in lignes_ok if l["horizon"] == h and l["dqp"] is not None]
        valeurs_par_horizon.append(valeurs)
        xs = [_horizon_en_minutes(h)] * len(valeurs)
        ax_disp.scatter(xs, valeurs, alpha=0.3, s=10, color="#1F618D", zorder=2)
    if any(valeurs_par_horizon):
        ax_disp.boxplot(
            valeurs_par_horizon, positions=positions, widths=largeur_boite,
            showfliers=False, patch_artist=True, zorder=3,
            boxprops=dict(facecolor=(0.682, 0.839, 0.945, 0.20), edgecolor="#154360", linewidth=1.2),
            medianprops=dict(color="#C0392B", linewidth=1.8),
            whiskerprops=dict(color="#154360", linewidth=1.2),
            capprops=dict(color="#154360", linewidth=1.2),
        )
    ax_disp.set_xticks([_horizon_en_minutes(h) for h in horizons])
    ax_disp.set_xticklabels(horizons, rotation=45, ha="right", fontsize=7)
    ax_disp.set_ylabel("|dQP| (%)", fontsize=8)
    ax_disp.set_title("Dispersion |dQP| par horizon", fontsize=9)
    ax_disp.grid(True, alpha=0.3)
    return fig


# ══════════════════════════════════════════════════════════════════════════════════
# Onglet 3 — Détail par crue
# ══════════════════════════════════════════════════════════════════════════════════

def _feuille_detail_par_crue(ws, lignes, infos_crues, meilleur, meilleur_combinaison_id,
                               paths, app, conn):
    _entete(ws, ENTETES_DETAIL, [26, 18, 16, 12, 12, 12, 12, 12, 12, 10, 20, 46])
    for l in lignes:
        info = infos_crues.get(l["crue_date"], {})
        # Date/heure en vraie valeur datetime (pas seulement dans le libellé texte
        # "#N - date") : triable/filtrable nativement dans Excel, comme l'était la
        # colonne "Date crue" de l'ancien export à 2 onglets — demandé explicitement
        # pour ne rien perdre de l'existant plutôt que de dupliquer tout un onglet
        # quasi identique.
        date_deb = info.get("date_deb") if info else None
        if date_deb is None:
            from datetime import datetime as _dt
            date_deb = _dt.fromisoformat(l["crue_date"])
        ws.append((
            _libelle_crue(l["crue_date"], info) if info else l["crue_date"],
            date_deb,
            l["horizon"], l["seuil_c1"], l["methode"], l["statut_crue"],
            l["dqp"], l["dtp"], l["ve"], l["kge"], l["suspects"] or "",
            # "Note" : erreur d'échec, OU précision sur un dQP/dTP recalculé depuis les
            # séries (pas extrait du PDF) ou toujours manquant malgré ce repli — voir
            # modules.run_orchestrator._recalculer_dqp_dtp. Jamais vide silencieusement
            # pour une valeur manquante (demandé explicitement par l'utilisateur).
            l.get("note") or "",
        ))

    ligne_libre = ws.max_row + 2
    seuils_q = app.config_data.get("seuils_q", {})
    if meilleur is None:
        ws.cell(row=ligne_libre, column=1,
                value="Aucune combinaison avec score exploitable — pas de graphique par crue.")
        return

    ws.cell(row=ligne_libre, column=1,
            value=f"Graphiques par crue — combinaison optimale : {meilleur.horizon} / "
                  f"seuil {meilleur.seuil_c1:.2f} / {meilleur.methode} "
                  f"(score {meilleur.score:.3f})").font = Font(bold=True, size=12)
    ligne_libre += 2

    for iso, info in sorted(infos_crues.items(),
                             key=lambda kv: (kv[1]["num_evt"] is None, kv[1]["num_evt"] or 0)):
        serie = info.get("serie")
        if not serie or info.get("code_pdt") is None:
            continue
        serie_sim = (results_store.charger_serie(conn, meilleur_combinaison_id, iso, "sim")
                     if meilleur_combinaison_id is not None else [])
        fig = _figure_detail_crue(_libelle_crue(iso, info), serie, serie_sim, meilleur, seuils_q)
        img = _fig_to_image(fig)
        ws.add_image(img, f"A{ligne_libre}")
        ligne_libre += 20  # hauteur approx. d'une image (dpi/figsize choisis plus bas) + marge


def _figure_detail_crue(libelle_crue, serie, serie_sim, meilleur, seuils_q):
    fig = Figure(figsize=(9.5, 3.6), dpi=100)
    ax = fig.add_subplot(1, 1, 1)
    ax_pluie = ax.twinx()
    ax_pluie.set_zorder(ax.get_zorder() - 1)
    ax.patch.set_visible(False)

    dates_obs = [p[0] for p in serie]
    qobs = [p[2] for p in serie]
    ax.plot(dates_obs, qobs, color=_COULEUR_OBS, lw=1.6, label="Q observé")
    toutes_valeurs = list(qobs)

    if serie_sim:
        dates_sim = [p[0] for p in serie_sim]
        qsim = [p[1] for p in serie_sim]
        ax.plot(dates_sim, qsim, color=_PALETTE_COURBES[0], lw=1.3, ls="--",
                label=f"Sim {meilleur.horizon}/{meilleur.seuil_c1:.2f}/{meilleur.methode}")
        toutes_valeurs.extend(qsim)

    intervalle = _intervalle_minutes(serie)
    if intervalle:
        profondeurs = [p[1] for p in serie]
        largeur_jours = (intervalle / (24 * 60)) * 0.8
        ax_pluie.bar(dates_obs, profondeurs, width=largeur_jours, color="#5DADE2",
                     edgecolor="#2E86AB", linewidth=0.3, alpha=0.75, zorder=1)
        plafond = max(max(profondeurs, default=0) * 4, 1)
        ax_pluie.set_ylim(plafond, 0)
        ax_pluie.yaxis.set_label_position("right")  # voir ui.tab_dashboard._tracer() : pas garanti par twinx() seul
        ax_pluie.set_ylabel("Pluie (mm)", fontsize=7, color="#2E86AB", labelpad=10)
        ax_pluie.tick_params(axis="y", labelsize=6.5, colors="#2E86AB")

    for cle, libelle in LIBELLES_SEUILS_Q:
        val = seuils_q.get(cle)
        if val is None or (toutes_valeurs and val > max(toutes_valeurs) * 1.3):
            continue
        est_zt = cle.startswith("zt_")
        ax.axhline(val, color="#784212", lw=0.9 if est_zt else 1.1, ls=":" if est_zt else "-",
                    alpha=0.6)

    ax.set_ylabel("Débit (m³/s)", fontsize=8)
    ax.set_title(libelle_crue, fontsize=9, loc="left")
    ax.grid(True, alpha=0.3)
    ax.tick_params(axis="both", labelsize=7)
    fig.autofmt_xdate()
    lignes1, labels1 = ax.get_legend_handles_labels()
    ax.legend(lignes1, labels1, loc="upper right", fontsize=6.5)
    fig.subplots_adjust(bottom=0.22, right=0.9)
    return fig


# ══════════════════════════════════════════════════════════════════════════════════
# Onglet 4 — Sensibilité au seuil
# ══════════════════════════════════════════════════════════════════════════════════

def _feuille_sensibilite_seuil(ws, scores_valides):
    """Reproduction statique de Dashboard > Sensibilité au seuil : contrairement à la
    version interactive (sélection d'un sous-ensemble d'horizons/méthodes), l'export
    montre TOUJOURS tous les horizons et les 2 méthodes ensemble — même principe que
    Vue synthèse/Vue 3D, qui n'exposent pas non plus l'état d'un filtre interactif."""
    _entete(ws, ENTETES_SENSIBILITE, [16, 12, 12, 18])
    for s in sorted(scores_valides, key=lambda s: (_horizon_en_minutes(s.horizon), s.seuil_c1, s.methode)):
        ws.append((s.horizon, s.seuil_c1, s.methode, round(s.score, 4)))

    if not scores_valides:
        return
    ligne_libre = ws.max_row + 2
    fig = _figure_sensibilite_seuil(scores_valides)
    img = _fig_to_image(fig)
    ws.add_image(img, f"A{ligne_libre}")


def _figure_sensibilite_seuil(scores_valides):
    fig = Figure(figsize=(10, 4.5), dpi=110)
    ax = fig.add_subplot(1, 1, 1)
    lignes_par_methode = {"T": "-", "R": "--"}
    horizons = sorted({s.horizon for s in scores_valides}, key=_horizon_en_minutes)
    for i, horizon in enumerate(horizons):
        couleur = _PALETTE_COURBES[i % len(_PALETTE_COURBES)]
        for methode, style in lignes_par_methode.items():
            scores_hm = sorted((s for s in scores_valides if s.horizon == horizon and s.methode == methode),
                                key=lambda s: s.seuil_c1)
            if not scores_hm:
                continue
            ax.plot([s.seuil_c1 for s in scores_hm], [s.score for s in scores_hm],
                    marker="o", markersize=3.5, lw=1.3, color=couleur, ls=style,
                    label=f"{horizon} ({methode})")
    ax.set_xlabel("Seuil de calage SeuilC1 (m³/s)", fontsize=8)
    ax.set_ylabel("Score composite (0=meilleur)", fontsize=8)
    ax.set_title("Sensibilité du score composite au seuil de calage — tous horizons/méthodes",
                 fontsize=9)
    ax.grid(True, alpha=0.3)
    ax.tick_params(labelsize=7)
    ax.legend(loc="best", fontsize=6, ncol=3 if len(horizons) > 4 else 2)
    fig.subplots_adjust(bottom=0.15, right=0.98)
    return fig


# ══════════════════════════════════════════════════════════════════════════════════
# Onglet 5 — Vue 3D
# ══════════════════════════════════════════════════════════════════════════════════

def _feuille_vue_3d(ws, scores_valides, meilleur):
    _entete(ws, ENTETES_VUE3D, [16, 12, 12, 18, 18, 18, 18, 18, 18])
    scores_tries = sorted(scores_valides, key=lambda s: s.score)
    for s in scores_tries:
        delta = s.score - meilleur.score if meilleur is not None else None
        ws.append((
            s.horizon, s.seuil_c1, s.methode, round(s.score, 4),
            round(delta, 4) if delta is not None else None,
            round(s.moyennes_erreur.get("dqp"), 2) if s.moyennes_erreur.get("dqp") is not None else None,
            round(s.moyennes_erreur.get("dtp"), 2) if s.moyennes_erreur.get("dtp") is not None else None,
            round(s.moyennes_erreur.get("ve"), 2) if s.moyennes_erreur.get("ve") is not None else None,
            round(s.moyennes_erreur.get("kge"), 4) if s.moyennes_erreur.get("kge") is not None else None,
        ))

    if not scores_valides or meilleur is None:
        return
    ligne_libre = ws.max_row + 2
    fig = _figure_vue_3d(scores_tries, meilleur)
    img = _fig_to_image(fig)
    ws.add_image(img, f"A{ligne_libre}")


def _figure_vue_3d(scores_tries, meilleur):
    from matplotlib import colormaps

    fig = Figure(figsize=(11.5, 5), dpi=115)
    ax = fig.add_subplot(1, 2, 1, projection="3d")
    ax_classement = fig.add_subplot(1, 2, 2)

    xs = [_horizon_en_minutes(s.horizon) for s in scores_tries]
    ys = [s.seuil_c1 for s in scores_tries]
    zs = [s.score for s in scores_tries]
    for methode, marqueur in _MARQUEURS_METHODE.items():
        indices = [i for i, s in enumerate(scores_tries) if s.methode == methode]
        if not indices:
            continue
        nuage = ax.scatter([xs[i] for i in indices], [ys[i] for i in indices],
                            [zs[i] for i in indices], c=[zs[i] for i in indices],
                            cmap="RdYlGn_r", vmin=0, vmax=1, marker=marqueur, s=55,
                            edgecolors="#333333", linewidths=0.5, label=f"Méthode {methode}")
    fig.colorbar(nuage, ax=ax, shrink=0.6, pad=0.1, label="Score composite (0=meilleur)")
    ax.scatter([_horizon_en_minutes(meilleur.horizon)], [meilleur.seuil_c1], [meilleur.score],
               marker="*", s=300, color="gold", edgecolors="#333333", linewidths=0.8,
               label="Meilleure combinaison")
    horizons_uniques = sorted({s.horizon for s in scores_tries}, key=_horizon_en_minutes)
    ax.set_xticks([_horizon_en_minutes(h) for h in horizons_uniques])
    ax.set_xticklabels(horizons_uniques, rotation=30, ha="right", fontsize=6)
    ax.set_xlabel("Horizon", labelpad=12, fontsize=7)
    ax.set_ylabel("Seuil de calage (m³/s)", labelpad=6, fontsize=7)
    ax.set_zlabel("Score composite", fontsize=7)
    ax.legend(loc="upper left", bbox_to_anchor=(-0.55, 1.0), fontsize=6.5)
    fig.subplots_adjust(left=0.16, wspace=0.4)

    cmap = colormaps["RdYlGn_r"]
    NB_MAX = 20
    scores_affiches = scores_tries[:NB_MAX]
    libelles = [f"{s.horizon} / {s.seuil_c1:.2f} / {s.methode}" for s in scores_affiches]
    deltas = [s.score - meilleur.score for s in scores_affiches]
    positions_y = list(range(len(scores_affiches)))[::-1]
    couleurs_barres = [cmap(s.score) for s in scores_affiches]
    ax_classement.barh(positions_y, deltas, color=couleurs_barres, edgecolor="#333333",
                        linewidth=0.5, height=0.7, zorder=2)
    ax_classement.scatter([0], [positions_y[0]], marker="*", s=180, color="gold",
                            edgecolors="#333333", linewidths=0.7, zorder=5)
    ax_classement.set_yticks(positions_y)
    ax_classement.set_yticklabels(libelles, fontsize=6.5)
    ax_classement.set_xlabel("Écart au meilleur (Δ)", fontsize=8)
    ax_classement.set_title(
        "Classement" + (f" (20 premières sur {len(scores_tries)})" if len(scores_tries) > NB_MAX else ""),
        fontsize=9)
    ax_classement.grid(True, axis="x", alpha=0.3)
    ax_classement.axvline(0, color="#333333", lw=0.8)
    return fig


# ══════════════════════════════════════════════════════════════════════════════════
# Onglet 6 — Variation selon le nombre de crues
# ══════════════════════════════════════════════════════════════════════════════════

def _points_variation_crues(lignes, infos_crues, poids, asymetrie_dtp):
    """Même logique que ui.tab_dashboard._build_variation_crues (voir ce module pour le
    détail) : pour N croissant de 3 au nombre total de crues avec Qmax connu (classées
    Qmax décroissant), la combinaison gagnante du score composite sur les N crues les
    plus fortes. Calculée sur TOUTES les crues réussies, volontairement indépendamment
    de la sélection "Crues incluses dans le score" (même choix que dans le Dashboard).
    Retourne une liste de (n, ScoreCombinaison gagnant à ce n)."""
    crues_avec_qmax = [(iso, info["qmax_obs"]) for iso, info in infos_crues.items()
                         if info.get("qmax_obs") is not None]
    if len(crues_avec_qmax) < 3:
        return []
    isos_ordre = [iso for iso, _q in sorted(crues_avec_qmax, key=lambda t: t[1], reverse=True)]

    lignes_success = [l for l in lignes if l["statut_crue"] == "success"]
    points = []
    for n in range(3, len(isos_ordre) + 1):
        isos_n = set(isos_ordre[:n])
        lignes_n = [l for l in lignes_success if l["crue_date"] in isos_n]
        if not lignes_n:
            continue
        scores_n = [s for s in calculer_scores(lignes_n, poids=poids, asymetrie_dtp=asymetrie_dtp)
                    if s.score is not None]
        if not scores_n:
            continue
        points.append((n, min(scores_n, key=lambda s: s.score)))
    return points


def _figure_variation_crues(points):
    """Reproduit le graphique de l'onglet Dashboard "Variation selon le nb de crues" :
    KGE moyen (brut, non normalisé) de la combinaison gagnante à chaque N, fond
    dégradé rouge/vert, lignes verticales + étiquettes aux bascules de combinaison
    gagnante. Pas d'icône ni de sélection interactive ici (statique, pour un export)."""
    ns = [n for n, _s in points]
    kges = [s.moyennes_erreur.get("kge") for _n, s in points]
    libelles_combo = [f"{s.horizon}/{s.seuil_c1:.2f}/{s.methode}" for _n, s in points]

    fig = Figure(figsize=(11, 4.6), dpi=100)
    ax = fig.add_subplot(1, 1, 1)
    ax.plot(ns, kges, color="#1F618D", lw=1.6, marker="o", markersize=3.5, zorder=3)
    ax.set_xlabel("N (crues les plus fortes retenues, Qmax décroissant)")
    ax.set_ylabel("KGE moyen — combinaison gagnante (brut, non normalisé)")
    ax.set_title("Stabilité de la combinaison optimale selon le nombre de crues retenues", fontsize=9)
    ax.grid(True, alpha=0.3)

    # Fond dégradé rouge (bas) -> vert (haut), calé sur les limites réellement
    # affichées de l'axe Y — mêmes principes que ui/tab_dashboard.py.
    xlim, ylim = ax.get_xlim(), ax.get_ylim()
    degrade = np.linspace(0, 1, 256).reshape(-1, 1)
    ax.imshow(degrade, extent=(*xlim, *ylim), aspect="auto", cmap="RdYlGn",
               alpha=0.15, zorder=0, origin="lower")
    ax.set_xlim(xlim)
    ax.set_ylim(ylim)

    precedent = None
    for i, (n, lib) in enumerate(zip(ns, libelles_combo)):
        if lib != precedent:
            ax.axvline(n, color="#7B7B7B", lw=0.7, ls="--", alpha=0.6, zorder=1)
            ax.annotate(lib, xy=(n, kges[i]), xytext=(4, 6), textcoords="offset points",
                        fontsize=6.5, rotation=90, va="bottom", ha="left", color="#333333")
            precedent = lib

    fig.subplots_adjust(bottom=0.14)
    return fig


def _feuille_variation_crues(ws, points, nb_crues_disponibles, libelle_profil):
    ws.append((f"Combinaison optimale par nombre de crues (N) — pondération : {libelle_profil}",))
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(("Crues classées par Qmax décroissant (les plus fortes en premier). "
                "Indépendant de la sélection \"Crues incluses dans le score\" (voir onglet "
                "Paramétrage) : ici N grandit automatiquement des crues les plus fortes "
                "jusqu'à la totalité disponible, pour observer si la combinaison optimale "
                "reste stable.",))
    ws.append(("⚠ Le score normalisé n'est PAS comparable d'un N à l'autre (min-max recalé "
                "sur le sous-ensemble de chaque N) — utile seulement pour désigner le gagnant "
                "à ce N précis. Le KGE moyen (brut), lui, est directement comparable d'un N à "
                "l'autre : privilégiez les zones stables (peu de variation) plutôt qu'un pic "
                "isolé, souvent obtenu à petit N et statistiquement fragile.",))
    ws.append(())

    if not points:
        ws.append((f"Pas assez de crues avec Qmax connu parmi les {nb_crues_disponibles} "
                    "testées pour cette analyse (minimum 3 requis, via CRITERES_PERF.DAT).",))
        return

    ws.append(ENTETES_VARIATION)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for n, s in points:
        ws.append((
            n, f"{s.horizon}/{s.seuil_c1:.2f}/{s.methode}", round(s.score, 4),
            round(s.moyennes_erreur.get("kge"), 3) if s.moyennes_erreur.get("kge") is not None else None,
            round(s.moyennes_erreur.get("dqp"), 2) if s.moyennes_erreur.get("dqp") is not None else None,
            round(s.moyennes_erreur.get("dtp"), 2) if s.moyennes_erreur.get("dtp") is not None else None,
        ))
    _ajuster_largeurs(ws, [30, 26, 22, 18, 22, 22])

    fig = _figure_variation_crues(points)
    img = _fig_to_image(fig)
    ws.add_image(img, f"A{ws.max_row + 2}")


# ══════════════════════════════════════════════════════════════════════════════════
# Onglet 7 — Analyse crues affl. (contribution des stations affluentes, sans calage)
# ══════════════════════════════════════════════════════════════════════════════════

def _barycentre_pluie(serie, date_limite=None):
    """Horodatage barycentrique de la pluie d'une série (date, pobs, qobs) — moyenne
    des horodatages pondérée par la lame Pobs de chaque pas de temps, None si la série
    ne contient aucune pluie. `date_limite` (typiquement le Qmax de l'exutoire) :
    si fourni, seules les pluies à cette date ou avant sont prises en compte — une
    pluie tombée APRÈS le pic n'a pas pu y contribuer. Copie de
    ui.tab_analyse_affluents._barycentre_pluie (calcul pur, modules/ ne doit pas
    dépendre de ui/ — voir l'en-tête de ce fichier)."""
    points = [(d, p) for d, p, _q in serie
              if p is not None and p > 0 and (date_limite is None or d <= date_limite)]
    if not points:
        return None
    d0 = points[0][0]
    poids_total = sum(p for _d, p in points)
    if poids_total <= 0:
        return None
    t_bary_s = sum((d - d0).total_seconds() * p for d, p in points) / poids_total
    return d0 + timedelta(seconds=t_bary_s)


def _tr_percentiles(paths, code_pdt, dates_iso_selectionnees):
    """P10/P50/P90 (minutes) du temps de réponse (Tr = Qmax exutoire − barycentre
    pluie) sur les crues actuellement sélectionnées pour la campagne
    (app.config_data["crues_selectionnees"]) — même périmètre que
    ui.tab_analyse_affluents._tr_crues_selectionnees. None si moins de 2 mesures
    exploitables (résultat trop instable pour être affiché)."""
    if not dates_iso_selectionnees:
        return None
    try:
        evenements = parse_criteres_perf(paths.criteres_perf_dat(code_pdt))
    except (FileNotFoundError, CriteresPerfError):
        return None
    trs = []
    for evt in evenements:
        if evt.date_deb.isoformat() not in dates_iso_selectionnees:
            continue
        chemin = os.path.join(paths.evenements_dir(code_pdt),
                               f"{paths.code_site}-EV{evt.num_evt:04d}.DAT")
        try:
            serie = parse_evenement_serie(chemin)
        except (FileNotFoundError, CriteresPerfError):
            continue
        _qmax, date_qmax = affluents.qmax_et_horodatage([(p[0], p[2]) for p in serie])
        if date_qmax is None:
            continue
        date_bary = _barycentre_pluie(serie, date_limite=date_qmax)
        if date_bary is None:
            continue
        trs.append((date_qmax - date_bary).total_seconds() / 60)
    if len(trs) < 2:
        return None
    return tuple(np.percentile(trs, [10, 50, 90]))


def _pas_de_temps_analyse_affluents(app):
    """Résout le pas de temps à utiliser : le dernier sélectionné dans l'outil
    (partagé entre Dashboard > Détail par crue, Crues et Analyse crues affl. — voir
    ui.widgets_common.libelle_dernier_pdt, dont ce n'est ici qu'une lecture directe de
    la clé de config qu'elle gère), ou le premier pas de temps configuré à défaut.
    Cette feuille ne dépend d'AUCUN résultat de campagne (contrairement aux 6 autres
    feuilles de ce module) : "Analyse crues affl." est indépendant de tout calage."""
    pdt_list = app.config_data.get("parametrage", {}).get("pas_de_temps", [])
    if not pdt_list:
        return None
    dernier_code = app.config_data.get("parametrage", {}).get("dernier_pdt_selectionne")
    if dernier_code and any(p["code"] == dernier_code for p in pdt_list):
        return dernier_code
    return pdt_list[0]["code"]


def _lister_crues_debit(paths, code_pdt):
    """Événements de type crue (TypEvt=Q) détectés pour ce pas de temps, triés par n°
    d'événement croissant — même filtrage/tri que ui.tab_analyse_affluents (liste des
    crues indépendante de tout résultat de campagne)."""
    try:
        evenements = parse_criteres_perf(paths.criteres_perf_dat(code_pdt))
    except (FileNotFoundError, CriteresPerfError):
        return []
    return sorted((e for e in evenements if e.est_crue), key=lambda e: e.num_evt)


def _figure_crue_affluents(evt, code_pdt, app, paths, liste_affl, tr_percentiles):
    """Reproduit le graphique de ui.tab_analyse_affluents pour une crue donnée : Qobs
    exutoire, débits affluents, hyétogramme, seuils de vigilance, bandes de propagation
    par affluent, barycentre de la pluie + temps de réponse (Tr), et 2 camemberts
    (contribution au pic exutoire, surface de BV suivie).

    Simplification volontaire par rapport à la version interactive : les cases à
    cocher (seuils/bandes de propagation/camemberts) n'ont pas d'équivalent dans un
    export statique — tout est toujours affiché. Pas de survol à la souris ni de
    panneau de vignettes latéral (repris dans le tableau "Volumes transités" qui suit
    la figure dans la feuille).

    Retourne (figure, lignes_bilan) — lignes_bilan = liste de (nom, surface_km2,
    volume_m3, pct_volume, q_a_qmax_exutoire, pct_qmax), une ligne par station
    (exutoire en premier), pour alimenter le tableau de données sans recalculer deux
    fois les mêmes séries."""
    nom_exutoire = (app.config_data.get("station", {}).get("nom_station") or "").strip() \
        or "Station exutoire"
    label_exutoire = f"Q observé — {nom_exutoire} (exutoire)"
    surface_exutoire = app.config_data.get("station", {}).get("surface_bv_km2")

    fig = Figure(figsize=(13, 4.6), dpi=100)
    gs = fig.add_gridspec(1, 3, width_ratios=(3.3, 1, 1), wspace=0.5)
    ax = fig.add_subplot(gs[0, 0])
    ax_pluie = ax.twinx()
    ax_pluie.set_zorder(ax.get_zorder() - 1)
    ax.patch.set_visible(False)
    ax_pic = fig.add_subplot(gs[0, 1])
    ax_surface = fig.add_subplot(gs[0, 2])

    chemin_serie = os.path.join(paths.evenements_dir(code_pdt),
                                 f"{paths.code_site}-EV{evt.num_evt:04d}.DAT")
    try:
        serie_exutoire = parse_evenement_serie(chemin_serie)
    except (FileNotFoundError, CriteresPerfError):
        serie_exutoire = []

    lignes_bilan = []
    qmax_exutoire = volume_exutoire = date_qmax_exutoire = None
    y_max_visible = None

    if serie_exutoire:
        points_exutoire = [(p[0], p[2]) for p in serie_exutoire]
        ax.plot([d for d, _v in points_exutoire], [v for _d, v in points_exutoire],
                color=_COULEUR_OBS, lw=1.8, label=label_exutoire)
        qmax_exutoire, date_qmax_exutoire = affluents.qmax_et_horodatage(points_exutoire)
        volume_exutoire = affluents.volume_m3(points_exutoire)
        if qmax_exutoire is not None:
            ax.plot([date_qmax_exutoire], [qmax_exutoire], marker="o", markersize=6,
                    color=_COULEUR_OBS, markeredgecolor="white", markeredgewidth=0.7, zorder=10)
            # Contribue à 100 % de son propre débit — référence des % de contribution
            # des affluents ci-dessous (même colonne, même instant de référence).
            lignes_bilan.append((f"{nom_exutoire} (exutoire)", surface_exutoire,
                                  volume_exutoire, None, qmax_exutoire, 100.0))
            if qmax_exutoire > 0:
                y_max_visible = qmax_exutoire * 1.15
                ax.set_ylim(0, y_max_visible)

        if len(serie_exutoire) >= 2:
            intervalle_minutes = (serie_exutoire[1][0] - serie_exutoire[0][0]).total_seconds() / 60
            if intervalle_minutes > 0:
                dates_pluie = [p[0] for p in serie_exutoire]
                profondeurs = [p[1] for p in serie_exutoire]
                largeur_jours = (intervalle_minutes / (24 * 60)) * 0.8
                ax_pluie.bar(dates_pluie, profondeurs, width=largeur_jours, color="#5DADE2",
                             edgecolor="#2E86AB", linewidth=0.3, alpha=0.75, zorder=1,
                             label="Pluie de bassin (exutoire)")
                plafond = max(max(profondeurs, default=0) * 4, 1)
                ax_pluie.set_ylim(plafond, 0)
                ax_pluie.yaxis.set_label_position("right")
                ax_pluie.set_ylabel("Pluie (mm / pas de temps)", fontsize=7, color="#2E86AB",
                                     labelpad=12)
                ax_pluie.tick_params(axis="y", labelsize=6.5, colors="#2E86AB")

    contributions_pie = []
    surfaces_pie = []
    for i, a in enumerate(liste_affl):
        if not a.fichier:
            continue
        try:
            serie_a, _nb_ignorees = affluents.charger_serie_affluent(
                a.fichier, evt.date_deb, evt.date_fin)
        except (FileNotFoundError, ValueError):
            continue
        if not serie_a:
            continue
        couleur = a.couleur or _PALETTE_COURBES[i % len(_PALETTE_COURBES)]
        ax.plot([d for d, _v in serie_a], [v for _d, v in serie_a],
                color=couleur, lw=1.3, ls="--", label=a.nom)
        qmax_a, date_qmax_a = affluents.qmax_et_horodatage(serie_a)
        volume_a = affluents.volume_m3(serie_a)
        if qmax_a is not None:
            ax.plot([date_qmax_a], [qmax_a], marker="o", markersize=6, color=couleur,
                    markeredgecolor="white", markeredgewidth=0.7, zorder=10)
        pct_volume = (volume_a / volume_exutoire * 100
                      if volume_a is not None and volume_exutoire else None)

        # Q rétropropagé : PAS le Qmax propre de l'affluent, mais son débit au moment
        # où l'eau qu'il fournissait atteignait (en théorie) le pic de l'exutoire —
        # càd à (horodatage du Qmax exutoire − P50 de CET affluent).
        q_retropropage = None
        if date_qmax_exutoire is not None and a.p50_min is not None:
            date_lookup = date_qmax_exutoire - timedelta(minutes=a.p50_min)
            q_retropropage, _d = affluents.valeur_au_plus_proche(serie_a, date_lookup)
        pct_qmax = (q_retropropage / qmax_exutoire * 100
                    if q_retropropage is not None and qmax_exutoire else None)
        lignes_bilan.append((a.nom, a.surface_bv_km2, volume_a, pct_volume, q_retropropage, pct_qmax))
        if pct_qmax is not None and pct_qmax > 0:
            contributions_pie.append((a.nom, couleur, pct_qmax))
        if a.surface_bv_km2:
            surfaces_pie.append((a.nom, couleur, a.surface_bv_km2))

        # Bande de propagation P10-P90 (+ trait P50) à partir du pic de CET affluent,
        # propagé sur l'axe temporel de l'exutoire — sens INVERSE de la
        # rétropropagation ci-dessus (ici : pic affluent -> pic exutoire).
        date_p10, date_p50, date_p90 = affluents.bornes_bande_propagation(date_qmax_a, a)
        if date_p10 is not None and date_p90 is not None:
            ax.axvspan(date_p10, date_p90, color=couleur, alpha=0.12, zorder=0)
        if date_p50 is not None:
            ax.axvline(date_p50, color=couleur, lw=2.0, alpha=0.7, zorder=2)

    seuils = app.config_data.get("seuils_q", {})
    for cle, libelle, couleur in _LIBELLES_SEUILS_Q_COULEUR:
        val = seuils.get(cle)
        if val is None or (y_max_visible is not None and val > y_max_visible):
            continue
        est_zt = cle.startswith("zt_")
        ax.axhline(val, color=couleur, lw=1.0 if est_zt else 1.3,
                   ls=":" if est_zt else "-", alpha=0.85)
        ax.text(0.002, val, f" {libelle} {val:.0f} m³/s", va="bottom", fontsize=6,
                color=couleur, transform=ax.get_yaxis_transform())

    # Pas de titre sur le graphique lui-même : l'identification de la crue est déjà la
    # cellule Excel juste au-dessus de l'image (voir _feuille_analyse_affluents) — un
    # titre ferait doublon et se chevauche avec l'annotation du barycentre de la pluie,
    # posée elle aussi sur le bord haut du graphique (voir plus bas).
    ax.set_ylabel("Débit (m³/s)", fontsize=8)
    ax.grid(True, alpha=0.3)
    ax.tick_params(axis="both", labelsize=7)
    lignes_ax, labels_ax = ax.get_legend_handles_labels()
    lignes_pluie, labels_pluie = ax_pluie.get_legend_handles_labels()
    ax.legend(lignes_ax + lignes_pluie, labels_ax + labels_pluie, loc="upper right", fontsize=6)

    # Barycentre de la pluie (marqueur sur le bord haut) + Tr, et bande statistique du
    # Tr calculée sur les crues sélectionnées pour la campagne (voir _tr_percentiles).
    if serie_exutoire:
        date_bary = _barycentre_pluie(serie_exutoire, date_limite=date_qmax_exutoire)
        if date_bary is not None:
            ax.plot([date_bary], [1], marker="o", markersize=6, color="#0B1F4B",
                    markeredgecolor="white", markeredgewidth=0.6, zorder=15, clip_on=False,
                    transform=ax.get_xaxis_transform())
            texte_tr = ""
            if date_qmax_exutoire is not None:
                minutes_tr = round((date_qmax_exutoire - date_bary).total_seconds() / 60)
                signe = "-" if minutes_tr < 0 else ""
                h_tr, m_tr = divmod(abs(minutes_tr), 60)
                texte_tr = f" — Tr={signe}{h_tr}h{m_tr:02d}"
            ax.annotate(
                f"Barycentre pluie {date_bary:%d/%m %H:%M}{texte_tr}",
                xy=(mdates.date2num(date_bary), 1), xycoords=ax.get_xaxis_transform(),
                xytext=(0, 6), textcoords="offset points", ha="center", va="bottom",
                fontsize=6, color="#0B1F4B", clip_on=False)
            if tr_percentiles is not None:
                p10_tr, p50_tr, p90_tr = tr_percentiles
                date_p10_tr = date_bary + timedelta(minutes=p10_tr)
                date_p50_tr = date_bary + timedelta(minutes=p50_tr)
                date_p90_tr = date_bary + timedelta(minutes=p90_tr)
                ax.axvspan(date_p10_tr, date_p90_tr, color="#2E86AB", alpha=0.10, zorder=0)
                ax.axvline(date_p50_tr, color="#2E86AB", lw=1.0, alpha=0.8, zorder=2)

    # -- Camembert 1 : contribution au pic exutoire (% de Q rétropropagé) ------------
    if contributions_pie:
        total_pct = sum(p for _n, _c, p in contributions_pie)
        if total_pct > 100:
            facteur = 100 / total_pct
            parts = [(n, c, p * facteur) for n, c, p in contributions_pie]
        else:
            parts = list(contributions_pie)
            reste = 100 - total_pct
            if reste > 0.5:
                parts.append(("Écoulements locaux", _COULEUR_LOCAL, reste))
        ax_pic.pie([p for _n, _c, p in parts], colors=[c for _n, c, _p in parts],
                   autopct=lambda v: f"{v:.0f}%" if v >= 5 else "",
                   textprops={"fontsize": 6}, wedgeprops={"edgecolor": "white", "linewidth": 0.6})
        couleur_pct = "#C0392B" if total_pct > 100 else "black"
        poids_pct = "bold" if total_pct > 100 else "normal"
        ax_pic.set_title(f"Contribution au pic exutoire\n{total_pct:.0f} %", fontsize=6.5,
                         color=couleur_pct, fontweight=poids_pct)
    else:
        ax_pic.axis("off")
        ax_pic.text(0.5, 0.5, "Aucune contribution\ncalculable", ha="center", va="center",
                    fontsize=6.5, color="#888888", transform=ax_pic.transAxes)
        ax_pic.set_title("Contribution au pic exutoire", fontsize=6.5)

    # -- Camembert 2 : surface de BV suivie -------------------------------------------
    if surface_exutoire and surface_exutoire > 0:
        pct_titre = sum(s for _n, _c, s in surfaces_pie) / surface_exutoire * 100
        total_suivi = sum(s for _n, _c, s in surfaces_pie)
        parts = list(surfaces_pie)
        if total_suivi > surface_exutoire:
            facteur = surface_exutoire / total_suivi
            parts = [(n, c, s * facteur) for n, c, s in parts]
            total_suivi = surface_exutoire
        reste = surface_exutoire - total_suivi
        if reste > 0.5:
            parts.append(("Écoulements locaux", _COULEUR_LOCAL, reste))
        if parts:
            ax_surface.pie([s for _n, _c, s in parts], colors=[c for _n, c, _s in parts],
                           autopct=lambda v: f"{v:.0f}%" if v >= 5 else "",
                           textprops={"fontsize": 6}, wedgeprops={"edgecolor": "white", "linewidth": 0.6})
        else:
            ax_surface.axis("off")
        ax_surface.set_title(f"Surface de BV suivie\n{pct_titre:.0f} %", fontsize=6.5)
    else:
        ax_surface.axis("off")
        ax_surface.text(0.5, 0.5, "Surface exutoire\ninconnue", ha="center", va="center",
                        fontsize=6.5, color="#888888", transform=ax_surface.transAxes)
        ax_surface.set_title("Surface de BV suivie", fontsize=6.5)

    fig.autofmt_xdate()
    return fig, lignes_bilan


def _feuille_analyse_affluents(ws, app, paths):
    """Reproduit ui.tab_analyse_affluents : contribution des stations affluentes à
    l'hydrogramme observé à l'exutoire — INDÉPENDANT de tout calage GRP (contrairement
    aux 6 autres feuilles de ce classeur, pas de résultat de campagne ici). Tableau de
    configuration des affluents, puis pour chaque crue détectée (TypEvt=Q) : un
    tableau de bilan (volumes/contributions) suivi de sa vignette graphique."""
    ws.append(("Analyse crues affluentes — contribution des stations affluentes à "
               "l'hydrogramme observé à l'exutoire",))
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(("Indépendant de tout calage GRP : ne compare que des débits OBSERVÉS "
               "(exutoire et affluents), contrairement aux autres feuilles de ce classeur "
               "qui évaluent une campagne de calage. Reproduction statique de l'onglet "
               "\"Analyse crues affl.\" — seuils de vigilance, bandes de propagation et "
               "camemberts toujours affichés (les cases à cocher de la version interactive "
               "n'ont pas d'équivalent ici).",))
    ws.append(())

    code_station = app.config_data.get("station", {}).get("code_station")
    config_affl = affluents.config_pour_station(app.config_data, code_station)
    liste_affl = [affluents.affluent_depuis_dict(d) for d in config_affl.get("liste", [])]
    if not liste_affl:
        ws.append(("Aucune station affluente configurée (onglet Analyse crues affl. > "
                   "Stations affluentes et temps de propagation).",))
        return

    ws.append(("Stations affluentes configurées",))
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
    _entete(ws, ENTETES_AFFLUENTS_CONFIG, [24, 16, 16, 14, 14, 14, 50])
    for a in liste_affl:
        ws.append((
            a.nom, a.code_station or "—",
            round(a.surface_bv_km2, 1) if a.surface_bv_km2 is not None else "—",
            affluents.minutes_vers_hhmm(a.p10_min) or "—",
            affluents.minutes_vers_hhmm(a.p50_min) or "—",
            affluents.minutes_vers_hhmm(a.p90_min) or "—",
            a.fichier or "—",
        ))
    ws.append(())

    code_pdt = _pas_de_temps_analyse_affluents(app)
    if not code_pdt or paths is None:
        ws.append(("Pas de temps ou dossiers de travail non configurés — impossible de "
                   "lister les crues.",))
        return
    crues = _lister_crues_debit(paths, code_pdt)
    if not crues:
        ws.append((f"Aucune crue détectée pour le pas de temps utilisé ({code_pdt}).",))
        return

    dates_iso_selectionnees = set(app.config_data.get("crues_selectionnees", []))
    tr_percentiles = _tr_percentiles(paths, code_pdt, dates_iso_selectionnees)

    ligne_libre = ws.max_row + 2
    for evt in crues:
        ws.cell(row=ligne_libre, column=1,
                value=f"Crue #{evt.num_evt} — {evt.date_deb:%d/%m/%Y %H:%M} "
                      f"(Qmax {evt.qmax:.1f} m³/s le {evt.date_qmax:%d/%m %H:%M})"
                ).font = Font(bold=True, size=11)
        ligne_libre += 1

        fig, lignes_bilan = _figure_crue_affluents(evt, code_pdt, app, paths, liste_affl,
                                                     tr_percentiles)

        for j, libelle in enumerate(ENTETES_BILAN_AFFLUENTS, start=1):
            ws.cell(row=ligne_libre, column=j, value=libelle).font = Font(bold=True)
        ligne_libre += 1
        for nom, surface, volume, pct_volume, q_ref, pct_qmax in lignes_bilan:
            ws.cell(row=ligne_libre, column=1, value=nom)
            ws.cell(row=ligne_libre, column=2,
                    value=round(surface, 1) if surface is not None else "—")
            ws.cell(row=ligne_libre, column=3,
                    value=round(volume / 1e6, 3) if volume is not None else "—")
            ws.cell(row=ligne_libre, column=4,
                    value=round(pct_volume, 1) if pct_volume is not None else "—")
            ws.cell(row=ligne_libre, column=5,
                    value=round(q_ref, 1) if q_ref is not None else "—")
            ws.cell(row=ligne_libre, column=6,
                    value=round(pct_qmax, 1) if pct_qmax is not None else "—")
            ligne_libre += 1

        ligne_libre += 1
        img = _fig_to_image(fig)
        ws.add_image(img, f"A{ligne_libre}")
        ligne_libre += 24  # hauteur approx. de l'image + marge (même heuristique que
                            # _feuille_detail_par_crue, non mesurée précisément)

    _ajuster_largeurs(ws, [26, 18, 20, 18, 22, 18])


# ══════════════════════════════════════════════════════════════════════════════════

def exporter(chemin_xlsx, app, db_path=None):
    """Génère un classeur à 7 feuilles depuis les résultats actuellement en base et la
    configuration actuelle de l'outil (station, seuils, pondération). Lève une
    exception explicite si aucun résultat n'existe encore."""
    with results_store.db_session(db_path) as conn:
        lignes = [dict(r) for r in results_store.list_resultats_avec_combinaison(conn)]
        if not lignes:
            raise ValueError("Aucun résultat de campagne en base — lancez d'abord une campagne "
                              "(onglet Campagne) avant d'exporter.")
        couverture = results_store.resume_couverture(conn)

        # "lignes_ok" (filtrée aux crues incluses dans le score) alimente le calcul du
        # score ET la dispersion |dQP| de la Vue synthèse — même principe que le
        # Dashboard, pour que l'export reflète exactement ce qui y est affiché. Détail
        # par crue, plus bas, utilise "lignes" (NON filtrée) : ce tableau croisé montre
        # tous les résultats, indépendamment de la sélection de crues du score.
        app.config_data.setdefault("score", config_ponderation_par_defaut())
        poids, asymetrie_dtp, libelle_profil = resoudre_ponderation(app.config_data["score"])
        crues_incluses = app.config_data["score"].get("crues_incluses")
        lignes_ok = filtrer_par_crues([l for l in lignes if l["statut_crue"] == "success"],
                                        crues_incluses)
        scores = calculer_scores(lignes_ok, poids=poids, asymetrie_dtp=asymetrie_dtp)
        scores_valides = [s for s in scores if s.score is not None]
        meilleur = min(scores_valides, key=lambda s: s.score) if scores_valides else None
        meilleur_combinaison_id = None
        if meilleur is not None:
            for l in lignes_ok:
                if (l["horizon"], l["seuil_c1"], l["methode"]) == \
                        (meilleur.horizon, meilleur.seuil_c1, meilleur.methode):
                    meilleur_combinaison_id = l["combinaison_id"]
                    break

        paths, _manquants = construire_grp_paths(app)
        dates_iso = sorted({l["crue_date"] for l in lignes})
        infos_crues = _construire_infos_crues(paths, app, dates_iso)

        wb = Workbook()
        ws_param = wb.active
        ws_param.title = "Paramétrage"
        _feuille_parametrage(ws_param, app, couverture, infos_crues, poids, asymetrie_dtp,
                              libelle_profil, crues_incluses)

        _feuille_vue_synthese(wb.create_sheet("Vue synthèse"), lignes_ok, scores_valides, meilleur)

        _feuille_detail_par_crue(wb.create_sheet("Détail par crue"), lignes, infos_crues,
                                   meilleur, meilleur_combinaison_id, paths, app, conn)

        _feuille_sensibilite_seuil(wb.create_sheet("Sensibilité au seuil"), scores_valides)

        _feuille_vue_3d(wb.create_sheet("Vue 3D"), scores_valides, meilleur)

        points_variation = _points_variation_crues(lignes, infos_crues, poids, asymetrie_dtp)
        _feuille_variation_crues(wb.create_sheet("Variation selon le nb de crues"),
                                   points_variation, len(infos_crues), libelle_profil)

        _feuille_analyse_affluents(wb.create_sheet("Analyse crues affl."), app, paths)

        os.makedirs(os.path.dirname(chemin_xlsx) or ".", exist_ok=True)
        wb.save(chemin_xlsx)
        return chemin_xlsx
